from openpyxl import load_workbook
import os
import glob

FileList = glob.glob('*.xlsx')

cwd_path = os.path.dirname(os.path.abspath(__file__))


def file_finder():
    file_name = ''
    for file in FileList:
        if file != 'output.xlsx':
            file_name = file
    print("Working on file: ", file_name)
    return file_name


workBook = load_workbook(filename=file_finder())
workingSheet = workBook.active
workingSheet2 = workBook.create_sheet("Sheet_B")
workingSheet2.title = "Summary"


# To consolidate all main sector into a sorted list (alphabetically)
def main_sector_consolidator():
    result = []
    for each_cell in workingSheet["H"][1:]:
        if each_cell.value not in result:
            result.append(each_cell.value)
    return sorted(result)


# To consolidate all sub sector into a sorted list (alphabetically)
def sub_sector_consolidator():
    result = []
    for each_cell in workingSheet["I"][1:]:
        each_cell_input = str(each_cell.value).split(',')
        for detailed_sector in each_cell_input:
            clean_ds = detailed_sector.lstrip()
            if clean_ds not in result:
                result.append(clean_ds)
    return sorted(result)


# To arrange a dictionary {main sector : all exclusive sub sector that is only in main sector}
def sector_dict_arranger():
    result = {}
    sector_index = 7;
    sub_sector_index = 8;
    for each_sector in main_sector_consolidator():
        result[each_sector] = []
    for row in workingSheet.iter_rows(min_row=2, values_only=True):
        for each_sub_sector in str(row[sub_sector_index]).split(','):
            clean_ss = each_sub_sector.lstrip()
            if clean_ss not in result[row[sector_index]]:
                result[row[sector_index]].append(clean_ss)
    return result


# To produce the binary data:
#   1)  if an organisation is under a sub sector, return 1
#       else, return 0
#   2)  For each organisation, arrange its corresponding 1s and 0s into a list
#       according to the sorted sub_sector_consolidated_list
#   3)  Then map out the dictionary of {Organisation : list mention in 2)}
def binary_data():
    binary_dict = {}
    sub_sector_data = sub_sector_consolidator()

    for organisation in workingSheet["B"][1:]:
        original_binary = []
        for i in sub_sector_data:
            original_binary.append(0)
        binary_dict[organisation.value] = original_binary

    for row in workingSheet.iter_rows(min_row=2):
        organisation = row[1].value
        sub_sector = str(row[8].value).split(',')
        for each_sub_sector in sub_sector:
            clean_ss = each_sub_sector.lstrip()
            each_sub_sector_index = sub_sector_data.index(clean_ss)
            binary_dict[organisation][each_sub_sector_index] = 1

    return binary_dict


def sub_sector_write():
    sub_sector_list = sub_sector_consolidator()
    i = 0
    header_index = 11
    for each_sub_sector in sub_sector_list:
        workingSheet.cell(row=1, column=header_index).value = each_sub_sector
        header_index += 1

    binary_input = binary_data()
    row_index = 2
    for row in workingSheet.iter_rows(min_row=2):
        organisation = row[1].value
        input_data = binary_input[organisation]
        column_index = 11
        for each_data in input_data:
            workingSheet.cell(row=row_index, column=column_index).value = each_data
            column_index += 1
        row_index += 1
    return


def summary_write():
    summary_data = sector_dict_arranger()
    row_index = 2
    workingSheet2.cell(row=1, column=1).value = "Main Sector"
    workingSheet2.cell(row=1, column=2).value = "Exclusive Sub Sector"
    workingSheet2.cell(row=1, column=3).value = "All Sub Sector"

    for sector in summary_data.keys():
        workingSheet2.cell(row=row_index, column=1).value = sector
        workingSheet2.cell(row=row_index, column=2).value = str(summary_data[sector])
        row_index += 1

    sub_sector_list = sub_sector_consolidator()
    row_index = 2
    for sub_sector in sub_sector_list:
        workingSheet2.cell(row=row_index, column=3).value = sub_sector
        row_index += 1
    return


sub_sector_write()
summary_write()

workBook.save('output.xlsx')
